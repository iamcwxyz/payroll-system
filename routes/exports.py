from flask import Blueprint, request, session, make_response
from auth import login_required, role_required
import database
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

exports_bp = Blueprint('exports', __name__)

@exports_bp.route('/export/employees')
@login_required
@role_required(['Admin', 'HR'])
def export_employees():
    conn = database.get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT employee_id, name, department, position, salary_rate, role, status
        FROM employees 
        ORDER BY employee_id
    """)
    employees = c.fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Directory"
    
    # Add header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Employee ID", "Full Name", "Department", "Position", "Daily Rate (₱)", "Role", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add data
    for row, emp in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=emp['employee_id'])
        ws.cell(row=row, column=2, value=emp['name'])
        ws.cell(row=row, column=3, value=emp['department'])
        ws.cell(row=row, column=4, value=emp['position'])
        ws.cell(row=row, column=5, value=f"₱{emp['salary_rate']:.2f}")
        ws.cell(row=row, column=6, value=emp['role'])
        ws.cell(row=row, column=7, value=emp['status'])
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Employee_Directory_{timestamp}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@exports_bp.route('/export/payroll')
@login_required
@role_required(['Admin', 'HR'])
def export_payroll():
    period = request.args.get('period', '')
    
    conn = database.get_db_connection()
    c = conn.cursor()
    
    if period:
        c.execute("""
            SELECT e.employee_id, e.name, e.department, e.position, 
                   p.period, p.base_salary, p.overtime, p.deductions, p.bonuses, p.net_pay
            FROM payroll p
            JOIN employees e ON p.employee_ref = e.id
            WHERE p.period = ?
            ORDER BY e.employee_id
        """, (period,))
    else:
        c.execute("""
            SELECT e.employee_id, e.name, e.department, e.position, 
                   p.period, p.base_salary, p.overtime, p.deductions, p.bonuses, p.net_pay
            FROM payroll p
            JOIN employees e ON p.employee_ref = e.id
            ORDER BY p.period DESC, e.employee_id
        """)
    
    payroll_data = c.fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    # Add header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Employee ID", "Name", "Department", "Position", "Period", "Base Salary", "Overtime", "Deductions", "Bonuses", "Net Pay"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add data
    for row, record in enumerate(payroll_data, 2):
        ws.cell(row=row, column=1, value=record['employee_id'])
        ws.cell(row=row, column=2, value=record['name'])
        ws.cell(row=row, column=3, value=record['department'])
        ws.cell(row=row, column=4, value=record['position'])
        ws.cell(row=row, column=5, value=record['period'])
        ws.cell(row=row, column=6, value=f"₱{record['base_salary']:.2f}")
        ws.cell(row=row, column=7, value=f"₱{record['overtime']:.2f}")
        ws.cell(row=row, column=8, value=f"₱{record['deductions']:.2f}")
        ws.cell(row=row, column=9, value=f"₱{record['bonuses']:.2f}")
        ws.cell(row=row, column=10, value=f"₱{record['net_pay']:.2f}")
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    period_suffix = f"_{period}" if period else ""
    filename = f"Payroll_Report{period_suffix}_{timestamp}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response